# -*- coding: utf-8 -*-
"""
VoiceVox Voice Generator Tool
エクセルの台詞リストからVoiceVox TTSで音声ファイルを一括生成するツール
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import threading
import tempfile
import time
import re
import wave
import struct

import openpyxl
import requests
import pygame

# 設定ファイルのパス
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

# VoiceVoxのデフォルトURL
VOICEVOX_URL = "http://localhost:50021"

# 感情判定用のキーワード
EMOTION_KEYWORDS = {
    "あまあま": ["好き", "大好き", "愛してる", "嬉しい", "幸せ", "ありがとう", "素敵", "可愛い", "優しい", "♡", "♥", "にこ", "わーい", "やったー"],
    "ツンツン": ["べ、別に", "バカ", "ばか", "アホ", "あほ", "うるさい", "知らない", "嫌い", "ふん", "はぁ？", "なによ", "ちがう", "違う", "勘違い"],
    "セクシー": ["ふふ", "うふふ", "ねぇ", "ダメ", "だめ", "いけない", "秘密", "ひみつ", "誘", "触", "キス", "抱"],
    "ささやき": ["しー", "内緒", "ないしょ", "こっそり", "静かに", "小声", "ひそひそ"],
    "ヒソヒソ": ["しー", "内緒", "ないしょ", "こっそり", "静かに", "小声", "ひそひそ"],
    "怒り": ["怒", "許さない", "ゆるさない", "ふざけるな", "なんだと", "くそ", "クソ", "ちくしょう", "畜生", "殺", "死ね"],
    "悲しみ": ["悲しい", "寂しい", "さみしい", "辛い", "つらい", "泣", "涙", "ごめん", "すまない", "申し訳"],
    "喜び": ["嬉しい", "うれしい", "楽しい", "たのしい", "わーい", "やった", "最高", "すごい", "素晴らしい"],
}


class VoiceVoxAPI:
    """VoiceVox API連携クラス"""
    
    def __init__(self, base_url: str = VOICEVOX_URL):
        self.base_url = base_url
    
    def is_running(self) -> bool:
        """VoiceVoxが起動しているか確認"""
        try:
            response = requests.get(f"{self.base_url}/speakers", timeout=3)
            return response.status_code == 200
        except:
            return False
    
    def get_speakers(self) -> list:
        """話者一覧を取得"""
        try:
            response = requests.get(f"{self.base_url}/speakers")
            response.raise_for_status()
            return response.json()
        except Exception as e:
            raise Exception(f"話者一覧の取得に失敗しました: {e}")
    
    def get_speaker_styles(self) -> dict:
        """話者とスタイルの辞書を作成 {speaker_name: [(style_name, style_id), ...]}"""
        speakers = self.get_speakers()
        result = {}
        for speaker in speakers:
            name = speaker["name"]
            styles = [(style["name"], style["id"]) for style in speaker["styles"]]
            result[name] = styles
        return result
    
    def generate_audio_query(self, text: str, speaker_id: int) -> dict:
        """音声合成用のクエリを生成"""
        response = requests.post(
            f"{self.base_url}/audio_query",
            params={"text": text, "speaker": speaker_id}
        )
        response.raise_for_status()
        return response.json()
    
    def synthesize(self, audio_query: dict, speaker_id: int) -> bytes:
        """音声を合成"""
        response = requests.post(
            f"{self.base_url}/synthesis",
            params={"speaker": speaker_id},
            json=audio_query
        )
        response.raise_for_status()
        return response.content
    
    def generate_speech(self, text: str, speaker_id: int) -> bytes:
        """テキストから音声を生成（WAV形式）"""
        try:
            query = self.generate_audio_query(text, speaker_id)
            audio_data = self.synthesize(query, speaker_id)
            return audio_data
        except Exception as e:
            raise Exception(f"音声生成に失敗しました: {e}")


class EmotionAnalyzer:
    """台詞から感情を分析するクラス"""
    
    @staticmethod
    def analyze(text: str, available_styles: list) -> str:
        """
        台詞から最適なスタイルを判定
        available_styles: [(style_name, style_id), ...]
        """
        available_style_names = [s[0] for s in available_styles]
        
        # 各感情のスコアを計算
        scores = {}
        for emotion, keywords in EMOTION_KEYWORDS.items():
            score = 0
            for keyword in keywords:
                if keyword in text:
                    score += 1
            if score > 0:
                scores[emotion] = score
        
        # スコアが高い順にソート
        if scores:
            sorted_emotions = sorted(scores.items(), key=lambda x: x[1], reverse=True)
            for emotion, _ in sorted_emotions:
                # 利用可能なスタイルに含まれているか確認
                for style_name in available_style_names:
                    if emotion in style_name or emotion.lower() in style_name.lower():
                        return style_name
        
        # デフォルトは「ノーマル」または最初のスタイル
        for style_name in available_style_names:
            if "ノーマル" in style_name or "normal" in style_name.lower():
                return style_name
        
        return available_style_names[0] if available_style_names else "ノーマル"


class ExcelReader:
    """エクセルファイル読み込みクラス"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        self.sheet = None
        self.cached_data = None
    
    def get_sheet_names(self) -> list:
        """シート名一覧を取得"""
        return self.workbook.sheetnames
    
    def set_sheet(self, sheet_name: str):
        """使用するシートを設定し、データをキャッシュ"""
        self.sheet = self.workbook[sheet_name]
        self.cached_data = []
        for row in self.sheet.iter_rows(values_only=True):
            self.cached_data.append(row)
    
    def get_column_letters(self) -> list:
        """列のアルファベット一覧を取得"""
        if not self.sheet:
            return []
        max_col = self.sheet.max_column
        return [openpyxl.utils.get_column_letter(i) for i in range(1, max_col + 1)]
    
    def _column_index(self, column_letter: str) -> int:
        """列文字を0始まりのインデックスに変換"""
        return openpyxl.utils.column_index_from_string(column_letter) - 1
    
    def get_unique_values_in_column(self, column_letter: str, start_row: int = 2) -> list:
        """指定列のユニークな値を取得"""
        if not self.cached_data:
            return []
        
        col_idx = self._column_index(column_letter)
        values = set()
        
        for row_idx in range(start_row - 1, len(self.cached_data)):
            row = self.cached_data[row_idx]
            if col_idx < len(row) and row[col_idx]:
                values.add(str(row[col_idx]).strip())
        
        return sorted(list(values))
    
    def get_rows_for_character(self, char_column: str, character: str, 
                                dialogue_column: str, filename_column: str, 
                                start_row: int) -> list:
        """特定キャラクターの台詞とファイル名を取得"""
        if not self.cached_data:
            return []
        
        char_idx = self._column_index(char_column)
        dialogue_idx = self._column_index(dialogue_column)
        filename_idx = self._column_index(filename_column)
        
        rows = []
        for row_idx in range(start_row - 1, len(self.cached_data)):
            row = self.cached_data[row_idx]
            
            if char_idx >= len(row):
                continue
                
            char_value = row[char_idx]
            if char_value and str(char_value).strip() == character:
                dialogue = row[dialogue_idx] if dialogue_idx < len(row) else None
                filename = row[filename_idx] if filename_idx < len(row) else None
                if dialogue and filename:
                    rows.append({
                        "dialogue": str(dialogue).strip(),
                        "filename": str(filename).strip()
                    })
        return rows
    
    def close(self):
        self.workbook.close()


class AudioConverter:
    """音声変換クラス"""
    
    @staticmethod
    def convert_to_16bit_44100hz(input_data: bytes, output_path: str):
        """WAVデータを16bit 44100Hzに変換して保存"""
        # 一時ファイルに保存
        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
            tmp.write(input_data)
            tmp_path = tmp.name
        
        try:
            # WAVファイルを読み込み
            with wave.open(tmp_path, 'rb') as wav_in:
                n_channels = wav_in.getnchannels()
                sampwidth = wav_in.getsampwidth()
                framerate = wav_in.getframerate()
                n_frames = wav_in.getnframes()
                audio_data = wav_in.readframes(n_frames)
            
            # VoiceVoxは24000Hzで出力するので、44100Hzにリサンプリング
            # 簡易的な方法：pydubを使用
            from pydub import AudioSegment
            audio = AudioSegment.from_wav(tmp_path)
            audio = audio.set_frame_rate(44100).set_sample_width(2).set_channels(2)
            audio.export(output_path, format="wav")
        finally:
            os.unlink(tmp_path)


class VoiceGeneratorApp:
    """メインアプリケーションクラス"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("VoiceVox Voice Generator")
        self.root.geometry("900x850")
        self.root.resizable(True, True)
        
        # 変数の初期化
        self.excel_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.char_column = tk.StringVar()
        self.dialogue_column = tk.StringVar()
        self.filename_column = tk.StringVar()
        self.start_row = tk.StringVar(value="2")
        self.output_path = tk.StringVar()
        self.auto_emotion = tk.BooleanVar(value=True)
        
        self.excel_reader = None
        self.voicevox_api = None
        self.speaker_styles = {}  # {speaker_name: [(style_name, style_id), ...]}
        self.characters = []
        self.voice_combos = {}  # {character: (speaker_var, style_var)}
        
        # pygame初期化（音声再生用）
        pygame.mixer.init()
        
        # 設定を読み込み
        self.load_config()
        
        # UIを構築
        self.build_ui()
        
        # VoiceVox接続確認
        self.check_voicevox()
    
    def load_config(self):
        """設定ファイルを読み込み"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    # 必要に応じて設定を読み込み
            except:
                pass
    
    def save_config(self):
        """設定ファイルに保存"""
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({}, f)
        except:
            pass
    
    def build_ui(self):
        """UIを構築"""
        # メインフレーム（スクロール可能）
        canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=canvas.yview)
        main_frame = ttk.Frame(canvas, padding="10")
        
        main_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=main_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # ===== セクション1: VoiceVox接続状態 =====
        section1 = ttk.LabelFrame(main_frame, text="① VoiceVox接続状態", padding="10")
        section1.pack(fill=tk.X, pady=(0, 10))
        
        status_frame = ttk.Frame(section1)
        status_frame.pack(fill=tk.X)
        
        self.voicevox_status = ttk.Label(status_frame, text="確認中...", foreground="gray")
        self.voicevox_status.pack(side=tk.LEFT)
        
        ttk.Button(status_frame, text="再接続", command=self.check_voicevox).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(status_frame, text="※VoiceVoxを起動してから「再接続」を押してください").pack(side=tk.LEFT, padx=(10, 0))
        
        # ===== セクション2: エクセルファイル =====
        section2 = ttk.LabelFrame(main_frame, text="② エクセルファイルを選択", padding="10")
        section2.pack(fill=tk.X, pady=(0, 10))
        
        excel_frame = ttk.Frame(section2)
        excel_frame.pack(fill=tk.X)
        
        ttk.Entry(excel_frame, textvariable=self.excel_path, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(excel_frame, text="参照...", command=self.browse_excel).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(excel_frame, text="読み込み", command=self.load_excel).pack(side=tk.LEFT, padx=(5, 0))
        
        # シート選択
        sheet_frame = ttk.Frame(section2)
        sheet_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(sheet_frame, text="使用するシート:").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_name, width=30, state="readonly")
        self.sheet_combo.pack(side=tk.LEFT, padx=(5, 10))
        ttk.Button(sheet_frame, text="シートを選択", command=self.select_sheet).pack(side=tk.LEFT)
        
        # ===== セクション3: 列指定 =====
        section3 = ttk.LabelFrame(main_frame, text="③ 列と開始行を指定", padding="10")
        section3.pack(fill=tk.X, pady=(0, 10))
        
        row1 = ttk.Frame(section3)
        row1.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(row1, text="キャラクター名の列:", width=20, anchor=tk.W).pack(side=tk.LEFT)
        self.char_column_combo = ttk.Combobox(row1, textvariable=self.char_column, width=10, state="readonly")
        self.char_column_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        row2 = ttk.Frame(section3)
        row2.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(row2, text="台詞の列:", width=20, anchor=tk.W).pack(side=tk.LEFT)
        self.dialogue_column_combo = ttk.Combobox(row2, textvariable=self.dialogue_column, width=10, state="readonly")
        self.dialogue_column_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        row3 = ttk.Frame(section3)
        row3.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(row3, text="ファイル名の列:", width=20, anchor=tk.W).pack(side=tk.LEFT)
        self.filename_column_combo = ttk.Combobox(row3, textvariable=self.filename_column, width=10, state="readonly")
        self.filename_column_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        row4 = ttk.Frame(section3)
        row4.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(row4, text="データ開始行:", width=20, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Entry(row4, textvariable=self.start_row, width=10).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(row4, text="（ヘッダーが1行目なら2を入力）").pack(side=tk.LEFT, padx=(10, 0))
        
        ttk.Button(section3, text="キャラクター一覧を読み込み", command=self.load_characters).pack(pady=(10, 0))
        
        # ===== セクション4: キャラクター選択 =====
        section4 = ttk.LabelFrame(main_frame, text="④ 書き出すキャラクターを選択（Ctrlキーで複数選択）", padding="10")
        section4.pack(fill=tk.X, pady=(0, 10))
        
        self.char_listbox_frame = ttk.Frame(section4)
        self.char_listbox_frame.pack(fill=tk.X)
        
        self.char_listbox = tk.Listbox(self.char_listbox_frame, selectmode=tk.MULTIPLE, height=6)
        self.char_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        char_scrollbar = ttk.Scrollbar(self.char_listbox_frame, orient=tk.VERTICAL, command=self.char_listbox.yview)
        char_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.char_listbox.config(yscrollcommand=char_scrollbar.set)
        
        # ===== セクション5: ボイス割り当て =====
        section5 = ttk.LabelFrame(main_frame, text="⑤⑥⑦ キャラクターにVoiceVoxボイスを割り当て（プレビュー可能）", padding="10")
        section5.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 感情自動判定オプション
        emotion_frame = ttk.Frame(section5)
        emotion_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Checkbutton(emotion_frame, text="台詞の内容から感情（スタイル）を自動判定する", 
                        variable=self.auto_emotion).pack(side=tk.LEFT)
        
        ttk.Button(section5, text="選択したキャラクターのボイス設定を開始", 
                   command=self.setup_voice_assignment).pack(pady=(0, 10))
        
        voice_canvas = tk.Canvas(section5, height=150)
        voice_scrollbar = ttk.Scrollbar(section5, orient=tk.VERTICAL, command=voice_canvas.yview)
        self.voice_assign_frame = ttk.Frame(voice_canvas)
        
        self.voice_assign_frame.bind(
            "<Configure>",
            lambda e: voice_canvas.configure(scrollregion=voice_canvas.bbox("all"))
        )
        
        voice_canvas.create_window((0, 0), window=self.voice_assign_frame, anchor="nw")
        voice_canvas.configure(yscrollcommand=voice_scrollbar.set)
        
        voice_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        voice_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # ===== セクション6: 出力設定 =====
        section6 = ttk.LabelFrame(main_frame, text="⑧ 出力設定", padding="10")
        section6.pack(fill=tk.X, pady=(0, 10))
        
        output_frame = ttk.Frame(section6)
        output_frame.pack(fill=tk.X)
        
        ttk.Label(output_frame, text="出力先フォルダ:").pack(side=tk.LEFT)
        ttk.Entry(output_frame, textvariable=self.output_path, width=50).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        ttk.Button(output_frame, text="参照...", command=self.browse_output).pack(side=tk.LEFT)
        
        self.generate_btn = ttk.Button(section6, text="🎵 音声ファイルを生成", command=self.generate_voices)
        self.generate_btn.pack(pady=(10, 0))
        
        self.progress = ttk.Progressbar(section6, mode="determinate")
        self.progress.pack(fill=tk.X, pady=(10, 0))
        
        self.status_label = ttk.Label(section6, text="")
        self.status_label.pack()
    
    def check_voicevox(self):
        """VoiceVoxの接続を確認"""
        self.voicevox_api = VoiceVoxAPI()
        
        if self.voicevox_api.is_running():
            self.voicevox_status.config(text="✓ VoiceVox接続OK", foreground="green")
            try:
                self.speaker_styles = self.voicevox_api.get_speaker_styles()
            except Exception as e:
                self.voicevox_status.config(text=f"話者取得エラー: {e}", foreground="red")
        else:
            self.voicevox_status.config(text="✗ VoiceVoxが起動していません", foreground="red")
    
    def browse_excel(self):
        """エクセルファイルを選択"""
        path = filedialog.askopenfilename(
            title="エクセルファイルを選択",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.excel_path.set(path)
    
    def load_excel(self):
        """エクセルファイルを読み込み"""
        if not self.excel_path.get():
            messagebox.showerror("エラー", "エクセルファイルを選択してください")
            return
        
        self.status_label.config(text="エクセルファイルを読み込み中...")
        self.root.update()
        
        try:
            if self.excel_reader:
                self.excel_reader.close()
            
            self.excel_reader = ExcelReader(self.excel_path.get())
            sheet_names = self.excel_reader.get_sheet_names()
            
            self.sheet_combo["values"] = sheet_names
            if sheet_names:
                self.sheet_combo.current(0)
            
            self.char_column_combo["values"] = []
            self.dialogue_column_combo["values"] = []
            self.filename_column_combo["values"] = []
            self.char_column.set("")
            self.dialogue_column.set("")
            self.filename_column.set("")
            
            self.char_listbox.delete(0, tk.END)
            self.characters = []
            
            self.status_label.config(text="")
            messagebox.showinfo("成功", f"エクセルファイルを読み込みました\nシート数: {len(sheet_names)}")
        except Exception as e:
            self.status_label.config(text="")
            messagebox.showerror("エラー", f"読み込みに失敗しました: {e}")
    
    def select_sheet(self):
        """シートを選択して列情報を読み込み"""
        if not self.excel_reader:
            messagebox.showerror("エラー", "先にエクセルファイルを読み込んでください")
            return
        
        if not self.sheet_name.get():
            messagebox.showerror("エラー", "シートを選択してください")
            return
        
        self.status_label.config(text="シートを読み込み中... しばらくお待ちください")
        self.root.update()
        
        try:
            self.excel_reader.set_sheet(self.sheet_name.get())
            columns = self.excel_reader.get_column_letters()
            
            self.char_column_combo["values"] = columns
            self.dialogue_column_combo["values"] = columns
            self.filename_column_combo["values"] = columns
            
            if columns:
                self.char_column.set(columns[0])
                if len(columns) > 1:
                    self.dialogue_column.set(columns[1])
                if len(columns) > 2:
                    self.filename_column.set(columns[2])
            
            self.char_listbox.delete(0, tk.END)
            self.characters = []
            
            self.status_label.config(text="")
            
            row_count = len(self.excel_reader.cached_data) if self.excel_reader.cached_data else 0
            messagebox.showinfo("成功", f"シート「{self.sheet_name.get()}」を読み込みました\n行数: {row_count}行")
        except Exception as e:
            self.status_label.config(text="")
            messagebox.showerror("エラー", f"シートの読み込みに失敗しました: {e}")
    
    def load_characters(self):
        """キャラクター一覧を読み込み"""
        if not self.excel_reader:
            messagebox.showerror("エラー", "先にエクセルファイルを読み込んでください")
            return
        
        if not self.excel_reader.cached_data:
            messagebox.showerror("エラー", "先にシートを選択してください")
            return
        
        if not self.char_column.get():
            messagebox.showerror("エラー", "キャラクター名の列を選択してください")
            return
        
        try:
            start_row = int(self.start_row.get())
        except:
            start_row = 2
        
        self.status_label.config(text="キャラクター一覧を作成中...")
        self.root.update()
        
        try:
            self.characters = self.excel_reader.get_unique_values_in_column(
                self.char_column.get(), start_row
            )
            self.char_listbox.delete(0, tk.END)
            for char in self.characters:
                self.char_listbox.insert(tk.END, char)
            
            self.status_label.config(text="")
            messagebox.showinfo("成功", f"{len(self.characters)}人のキャラクターが見つかりました")
        except Exception as e:
            self.status_label.config(text="")
            messagebox.showerror("エラー", f"読み込みに失敗しました: {e}")
    
    def setup_voice_assignment(self):
        """選択したキャラクターのボイス割り当てUIを構築"""
        selected_indices = self.char_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("エラー", "書き出すキャラクターを選択してください")
            return
        
        if not self.speaker_styles:
            messagebox.showerror("エラー", "VoiceVoxが接続されていません。VoiceVoxを起動して「再接続」を押してください")
            return
        
        for widget in self.voice_assign_frame.winfo_children():
            widget.destroy()
        
        selected_chars = [self.characters[i] for i in selected_indices]
        speaker_names = list(self.speaker_styles.keys())
        
        self.voice_combos = {}
        
        for i, char in enumerate(selected_chars):
            row_frame = ttk.Frame(self.voice_assign_frame)
            row_frame.pack(fill=tk.X, pady=5)
            
            ttk.Label(row_frame, text=f"{char}:", width=15, anchor=tk.W).pack(side=tk.LEFT)
            
            # 話者選択
            speaker_var = tk.StringVar()
            speaker_combo = ttk.Combobox(row_frame, textvariable=speaker_var, values=speaker_names, width=15, state="readonly")
            speaker_combo.pack(side=tk.LEFT, padx=(5, 5))
            if speaker_names:
                speaker_combo.current(0)
            
            # スタイル選択
            style_var = tk.StringVar()
            style_combo = ttk.Combobox(row_frame, textvariable=style_var, width=12, state="readonly")
            style_combo.pack(side=tk.LEFT, padx=(5, 10))
            
            # 話者が変更されたらスタイルを更新
            def update_styles(event, sv=speaker_var, sc=style_combo, stv=style_var):
                speaker = sv.get()
                if speaker in self.speaker_styles:
                    styles = [s[0] for s in self.speaker_styles[speaker]]
                    sc["values"] = styles
                    if styles:
                        sc.current(0)
            
            speaker_combo.bind("<<ComboboxSelected>>", update_styles)
            
            # 初期スタイルを設定
            if speaker_names:
                first_speaker = speaker_names[0]
                styles = [s[0] for s in self.speaker_styles[first_speaker]]
                style_combo["values"] = styles
                if styles:
                    style_combo.current(0)
            
            self.voice_combos[char] = (speaker_var, style_var, style_combo)
            
            # プレビューボタン
            preview_btn = ttk.Button(row_frame, text="▶ プレビュー", 
                                     command=lambda c=char: self.preview_voice(c))
            preview_btn.pack(side=tk.LEFT)
        
        messagebox.showinfo("準備完了", f"{len(selected_chars)}人のキャラクターのボイス設定ができます")
    
    def get_style_id(self, speaker_name: str, style_name: str) -> int:
        """話者名とスタイル名からスタイルIDを取得"""
        if speaker_name in self.speaker_styles:
            for name, id in self.speaker_styles[speaker_name]:
                if name == style_name:
                    return id
        return 0
    
    def preview_voice(self, character: str):
        """選択したボイスでプレビュー再生（選択中のスタイルをそのまま使用）"""
        if not self.excel_reader:
            messagebox.showerror("エラー", "エクセルファイルを読み込んでください")
            return
        
        if character not in self.voice_combos:
            return
        
        speaker_var, style_var, style_combo = self.voice_combos[character]
        speaker_name = speaker_var.get()
        style_name = style_var.get()
        
        if not speaker_name or not style_name:
            messagebox.showerror("エラー", "話者とスタイルを選択してください")
            return
        
        try:
            start_row = int(self.start_row.get())
        except:
            start_row = 2
        
        rows = self.excel_reader.get_rows_for_character(
            self.char_column.get(), character,
            self.dialogue_column.get(), self.filename_column.get(),
            start_row
        )
        
        if not rows:
            messagebox.showerror("エラー", f"{character}の台詞が見つかりません")
            return
        
        first_dialogue = rows[0]["dialogue"]
        
        # プレビューでは選択中のスタイルをそのまま使用（自動判定しない）
        style_id = self.get_style_id(speaker_name, style_name)
        
        def generate_preview():
            tmp_path = None
            try:
                display_text = first_dialogue[:30] + "..." if len(first_dialogue) > 30 else first_dialogue
                self.status_label.config(text=f"プレビュー生成中: 「{display_text}」")
                wav_data = self.voicevox_api.generate_speech(first_dialogue, style_id)
                
                with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
                    tmp.write(wav_data)
                    tmp_path = tmp.name
                
                pygame.mixer.music.load(tmp_path)
                pygame.mixer.music.play()
                
                self.status_label.config(text=f"再生中: {character} - {speaker_name}（{style_name}）")
                
                while pygame.mixer.music.get_busy():
                    time.sleep(0.1)
                
                self.status_label.config(text="")
                
            except Exception as e:
                self.status_label.config(text=f"エラー: {str(e)[:50]}")
            finally:
                if tmp_path:
                    try:
                        time.sleep(0.3)
                        os.unlink(tmp_path)
                    except:
                        pass
        
        threading.Thread(target=generate_preview, daemon=True).start()
    
    def browse_output(self):
        """出力先フォルダを選択"""
        path = filedialog.askdirectory(title="出力先フォルダを選択")
        if path:
            self.output_path.set(path)
    
    def generate_voices(self):
        """音声ファイルを一括生成"""
        if not self.output_path.get():
            messagebox.showerror("エラー", "出力先フォルダを選択してください")
            return
        
        if not self.voice_combos:
            messagebox.showerror("エラー", "キャラクターのボイス設定を行ってください")
            return
        
        if not self.excel_reader:
            messagebox.showerror("エラー", "エクセルファイルを読み込んでください")
            return
        
        if not self.voicevox_api or not self.voicevox_api.is_running():
            messagebox.showerror("エラー", "VoiceVoxが起動していません")
            return
        
        os.makedirs(self.output_path.get(), exist_ok=True)
        
        try:
            start_row = int(self.start_row.get())
        except:
            start_row = 2
        
        # タスクを収集
        tasks = []
        for char, (speaker_var, style_var, style_combo) in self.voice_combos.items():
            speaker_name = speaker_var.get()
            base_style_name = style_var.get()
            
            rows = self.excel_reader.get_rows_for_character(
                self.char_column.get(), char,
                self.dialogue_column.get(), self.filename_column.get(),
                start_row
            )
            
            for row in rows:
                dialogue = row["dialogue"]
                
                # 感情自動判定
                if self.auto_emotion.get():
                    styles = self.speaker_styles.get(speaker_name, [])
                    style_name = EmotionAnalyzer.analyze(dialogue, styles)
                else:
                    style_name = base_style_name
                
                style_id = self.get_style_id(speaker_name, style_name)
                
                tasks.append({
                    "character": char,
                    "speaker": speaker_name,
                    "style": style_name,
                    "style_id": style_id,
                    "dialogue": dialogue,
                    "filename": row["filename"]
                })
        
        if not tasks:
            messagebox.showerror("エラー", "生成する台詞がありません")
            return
        
        if not messagebox.askyesno("確認", f"{len(tasks)}個の音声ファイルを生成しますか？"):
            return
        
        def generate_all():
            self.generate_btn.config(state=tk.DISABLED)
            self.progress["maximum"] = len(tasks)
            self.progress["value"] = 0
            
            success_count = 0
            error_count = 0
            
            for i, task in enumerate(tasks):
                try:
                    self.status_label.config(
                        text=f"生成中 ({i+1}/{len(tasks)}): {task['filename']}"
                    )
                    self.root.update()
                    
                    wav_data = self.voicevox_api.generate_speech(
                        task["dialogue"], task["style_id"]
                    )
                    
                    filename = task["filename"]
                    if not filename.lower().endswith(".wav"):
                        filename += ".wav"
                    
                    output_file = os.path.join(self.output_path.get(), filename)
                    
                    # 16bit 44100Hzに変換して保存
                    AudioConverter.convert_to_16bit_44100hz(wav_data, output_file)
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    print(f"Error generating {task['filename']}: {e}")
                
                self.progress["value"] = i + 1
                self.root.update()
            
            self.generate_btn.config(state=tk.NORMAL)
            self.status_label.config(text="")
            
            messagebox.showinfo(
                "完了",
                f"音声生成が完了しました\n成功: {success_count}件\nエラー: {error_count}件"
            )
        
        threading.Thread(target=generate_all, daemon=True).start()
    
    def run(self):
        """アプリケーションを実行"""
        self.root.mainloop()
        
        if self.excel_reader:
            self.excel_reader.close()
        pygame.mixer.quit()


if __name__ == "__main__":
    app = VoiceGeneratorApp()
    app.run()
